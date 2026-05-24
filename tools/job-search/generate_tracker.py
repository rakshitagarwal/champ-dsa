#!/usr/bin/env python3
"""Generate job_search_tracker.xlsx — cold outreach job search workbook."""

from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = Path(__file__).resolve().parent
OUTPUT = DIR / "job_search_tracker.xlsx"
COMPANIES_FILE = DIR / "companies.json"

# Styling
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=11)
LINK_FONT = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
MONO_FONT = Font(name="Consolas", size=10)
EMAIL_BLOCK_FILL = PatternFill("solid", fgColor="F3F4F6")
SAMPLE_FILL = PatternFill("solid", fgColor="FEF08A")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

TIER_FILLS = {
    1: PatternFill("solid", fgColor="FFF9C4"),
    2: PatternFill("solid", fgColor="DBEAFE"),
    3: PatternFill("solid", fgColor="DCFCE7"),
    4: PatternFill("solid", fgColor="FCE7F3"),
    5: PatternFill("solid", fgColor="E0E7FF"),
}

YES_NO_DASH = '"Yes,No,—"'


def linkedin_people_search(company: str) -> str:
    q = f"{company} recruiter OR talent acquisition"
    return f"https://www.linkedin.com/search/results/people/?keywords={quote(q)}"


def style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def set_hyperlink(cell, url: str, display: str) -> None:
    cell.value = display
    cell.hyperlink = url
    cell.font = LINK_FONT


def apply_body_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.border.left.style is None:
                cell.border = THIN_BORDER
            cell.font = cell.font if cell.font != LINK_FONT else LINK_FONT
            if cell.font.name != "Consolas" and cell != ws.cell(row=row, column=col):
                pass


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    lines = [
        ("Cold Outreach Job Search Tracker", BOLD_FONT),
        ("", BODY_FONT),
        (
            "Target: Software Engineer / Web Developer / MERN / Full stack · 3–6 years · "
            "Bangalore, Hyderabad, Pune, Delhi NCR, Remote India · Fintech, AI startups, SaaS, IT, startups",
            BODY_FONT,
        ),
        ("", BODY_FONT),
        ("HOW TO USE THIS WORKBOOK", BOLD_FONT),
        ("", BODY_FONT),
        ("Tab: Company Tracker", BOLD_FONT),
        (
            "Work through 50 curated companies by tier. Click “Search recruiters” to open a pre-built "
            "LinkedIn people search. Click “Careers page” to check open roles. Update Priority, Status, "
            "People Found, and Emails Sent as you progress.",
            BODY_FONT,
        ),
        ("", BODY_FONT),
        ("Tab: Outreach Tracker", BOLD_FONT),
        (
            "Log every contact you email. Fill First Name through Notes. Date Sent auto-calculates "
            "Days Since Sent. Track opens, replies, and follow-ups. Use the yellow sample row as a format guide.",
            BODY_FONT,
        ),
        ("", BODY_FONT),
        ("Tab: Email Template", BOLD_FONT),
        (
            "Copy Version A (recruiters/TA) or Version B (engineering managers). Replace "
            "{{First Name}}, {{Company}}, and {{Hook}}. Send follow-up 5 days later if no reply.",
            BODY_FONT,
        ),
        ("", BODY_FONT),
        ("DAILY 30-MINUTE WORKFLOW", BOLD_FONT),
        ("", BODY_FONT),
        ("1. Pick 3–5 companies from Company Tracker (start Tier 1, then expand).", BODY_FONT),
        ("2. Open LinkedIn people search link → find 2 recruiters or TA per company.", BODY_FONT),
        ("3. Use free email finders (Hunter, Snov, Apollo, etc.) to get verified emails.", BODY_FONT),
        ("4. Write a one-line hook specific to that person or company.", BODY_FONT),
        ("5. Send via mail-merge (YAMM / GMass / Mailmeteor) — never BCC at scale.", BODY_FONT),
        ("6. Log each send in Outreach Tracker; follow up after 5 days if no reply.", BODY_FONT),
        ("", BODY_FONT),
        ("FREE EMAIL-FINDER STACK (approx. monthly free tiers)", BOLD_FONT),
        ("", BODY_FONT),
        ("• Apollo.io — ~50 credits/month on free plan", BODY_FONT),
        ("• Hunter.io — ~25 searches/month", BODY_FONT),
        ("• Snov.io — ~50 credits/month", BODY_FONT),
        ("• FindThatLead — limited free searches", BODY_FONT),
        ("• RocketReach — limited free lookups", BODY_FONT),
        ("", BODY_FONT),
        ("FREE MAIL-MERGE TOOLS (daily limits)", BOLD_FONT),
        ("", BODY_FONT),
        ("• YAMM (Yet Another Mail Merge) — ~50 emails/day on free Gmail", BODY_FONT),
        ("• GMass — free tier with daily send caps", BODY_FONT),
        ("• Mailmeteor — ~75 emails/day on free plan", BODY_FONT),
        ("", BODY_FONT),
        ("WARNINGS", BOLD_FONT),
        ("", BODY_FONT),
        ("• Never BCC large cold-email batches — use individual merge sends.", BODY_FONT),
        ("• Warm up a new inbox: ~10 emails/day for 3–4 days before scaling.", BODY_FONT),
        ("• No links or attachments in the first email (spam-filter risk).", BODY_FONT),
        ("• Send 9–11 AM in the recipient’s time zone when possible.", BODY_FONT),
        ("• If 0 replies after 30 sends, rewrite your hook and subject line.", BODY_FONT),
    ]

    for idx, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=idx, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = 18 if text else 8

    ws.row_dimensions[1].height = 28


def build_company_tracker(wb: Workbook, companies: list[dict]) -> None:
    ws = wb.create_sheet("Company Tracker")
    ws.sheet_view.showGridLines = False

    headers = [
        "#",
        "Company",
        "Tier",
        "HQ/Location",
        "Why Target",
        "LinkedIn People Search URL",
        "Careers Page",
        "Priority",
        "Status",
        "People Found",
        "Emails Sent",
        "Notes",
    ]
    col_widths = [5, 22, 28, 22, 42, 22, 18, 12, 14, 14, 14, 24]

    # Stats rows 1–3
    ws["A1"] = "Total companies"
    ws["A1"].font = BOLD_FONT
    ws["B1"] = "=COUNTA(B5:B54)"
    ws["B1"].font = BODY_FONT

    ws["A2"] = "Total people found"
    ws["A2"].font = BOLD_FONT
    ws["B2"] = "=SUM(J5:J54)"
    ws["B2"].font = BODY_FONT

    ws["A3"] = "Total emails sent"
    ws["A3"].font = BOLD_FONT
    ws["B3"] = "=SUM(K5:K54)"
    ws["B3"].font = BODY_FONT

    header_row = 4
    data_start = 5
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        ws.cell(row=header_row, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width
    style_header_row(ws, header_row, len(headers))

    priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    status_dv = DataValidation(
        type="list",
        formula1='"Not started,Researching,Contacts found,Outreach started,Replied,Closed"',
        allow_blank=True,
    )
    ws.add_data_validation(priority_dv)
    ws.add_data_validation(status_dv)

    for i, co in enumerate(companies):
        row = data_start + i
        tier_key = co["tier_key"]
        fill = TIER_FILLS.get(tier_key, PatternFill())

        ws.cell(row=row, column=1, value=i + 1).font = BODY_FONT
        ws.cell(row=row, column=2, value=co["company"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=co["tier"]).font = BODY_FONT
        ws.cell(row=row, column=4, value=co["hq"]).font = BODY_FONT
        ws.cell(row=row, column=5, value=co["why_target"]).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

        li_cell = ws.cell(row=row, column=6)
        set_hyperlink(li_cell, linkedin_people_search(co["company"]), "Search recruiters")

        careers_cell = ws.cell(row=row, column=7)
        set_hyperlink(careers_cell, co["careers_url"], "Careers page")

        ws.cell(row=row, column=8, value="Medium").font = BODY_FONT
        ws.cell(row=row, column=9, value="Not started").font = BODY_FONT
        ws.cell(row=row, column=10, value=0).font = BODY_FONT
        ws.cell(row=row, column=11, value=0).font = BODY_FONT
        ws.cell(row=row, column=12, value="").font = BODY_FONT

        priority_dv.add(ws.cell(row=row, column=8))
        status_dv.add(ws.cell(row=row, column=9))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            if col not in (6, 7):
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 5))

    ws.freeze_panes = "B5"


def build_outreach_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("Outreach Tracker")
    ws.sheet_view.showGridLines = False

    headers = [
        "#",
        "First Name",
        "Last Name",
        "Company",
        "Role/Title",
        "LinkedIn URL",
        "Email",
        "Email Source",
        "Hook",
        "Status",
        "Date Sent",
        "Days Since Sent",
        "Opened?",
        "Replied?",
        "Follow-up Sent?",
        "Outcome",
        "Notes",
    ]
    col_widths = [5, 14, 14, 18, 22, 28, 28, 14, 36, 12, 14, 16, 10, 10, 16, 14, 24]

    # Stats rows 1–4 (B3/B4 used for reply-rate formula per spec)
    ws["A1"] = "Total contacts"
    ws["A1"].font = BOLD_FONT
    ws["B1"] = "=COUNTA(B6:B106)"
    ws["B1"].font = BODY_FONT

    ws["A3"] = "Emails sent"
    ws["A3"].font = BOLD_FONT
    ws["B3"] = "=COUNTA(K6:K106)"
    ws["B3"].font = BODY_FONT

    ws["A4"] = "Replies"
    ws["A4"].font = BOLD_FONT
    ws["B4"] = '=COUNTIF(N6:N106,"Yes")'
    ws["B4"].font = BODY_FONT

    ws["D1"] = "Reply rate"
    ws["D1"].font = BOLD_FONT
    ws["E1"] = '=IF(B3=0,"",B4/B3)'
    ws["E1"].font = BODY_FONT
    ws["E1"].number_format = "0.0%"

    header_row = 5
    data_start = 6
    sample_row = 6

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        ws.cell(row=header_row, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = width
    style_header_row(ws, header_row, len(headers))

    yes_no_dv = DataValidation(type="list", formula1=YES_NO_DASH, allow_blank=True)
    ws.add_data_validation(yes_no_dv)

    # Sample row (yellow) — first data row
    sample_date = date.today() - timedelta(days=7)
    sample = {
        1: 1,
        2: "Priya",
        3: "Sharma",
        4: "Razorpay",
        5: "Senior Talent Acquisition Partner",
        6: "https://www.linkedin.com/in/example-priya-sharma",
        7: "priya.sharma@example.com",
        8: "Hunter.io",
        9: "Razorpay's payments stack at UPI scale is exactly the kind of fintech product work I want to grow into.",
        10: "Sent",
        11: sample_date,
        13: "No",
        14: "No",
        15: "No",
        16: "Pending",
        17: "Sample row — delete or overwrite when logging real contacts.",
    }

    for col, val in sample.items():
        cell = ws.cell(row=sample_row, column=col, value=val)
        cell.fill = SAMPLE_FILL
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if col == 6:
            set_hyperlink(cell, str(val), "LinkedIn profile")
            cell.fill = SAMPLE_FILL
        if col == 11:
            cell.number_format = "yyyy-mm-dd"

    ws.cell(row=sample_row, column=12, value=f'=IF(K{sample_row}="","",TODAY()-K{sample_row})')
    ws.cell(row=sample_row, column=12).fill = SAMPLE_FILL
    ws.cell(row=sample_row, column=12).font = BODY_FONT
    ws.cell(row=sample_row, column=12).border = THIN_BORDER

    for col in (13, 14, 15):
        yes_no_dv.add(ws.cell(row=sample_row, column=col))

    for row in range(data_start + 1, data_start + 101):
        ws.cell(row=row, column=1, value=row - sample_row).font = BODY_FONT
        ws.cell(row=row, column=12, value=f'=IF(K{row}="","",TODAY()-K{row})').font = BODY_FONT
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col == 11:
                cell.number_format = "yyyy-mm-dd"
            if col in (13, 14, 15):
                yes_no_dv.add(cell)
            if col != 12:
                cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 9, 17)))

    ws.freeze_panes = "B6"


def build_email_template(wb: Workbook) -> None:
    ws = wb.create_sheet("Email Template")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    version_a = """Subject: Full-stack engineer — quick question about {{Company}}

Hi {{First Name}},

{{Hook}}

I'm a full-stack developer (3–6 yrs) with MERN experience, currently exploring roles in Bangalore, Hyderabad, Pune, and remote India — especially fintech and SaaS.

Would you be open to a 15-minute chat if there's alignment on your open web engineering roles?

Best,
[Your Name]
[Phone] · [LinkedIn]"""

    version_b = """Subject: Advice on breaking into engineering at {{Company}}?

Hi {{First Name}},

{{Hook}}

I'm a full-stack developer with ~4 years on MERN and product web apps. I'm not asking for a referral yet — I'd genuinely value 10 minutes of your perspective on what strong mid-level engineers look like on your team and how you'd suggest approaching {{Company}}.

If you're open to it, I can share my background briefly and keep it short.

Thanks for considering,
[Your Name]
[LinkedIn]"""

    follow_up = """Subject: Re: quick follow-up

Hi {{First Name}},

Just bumping my note from last week — totally understand if you're swamped.

Still interested in learning whether there might be fit for a full-stack engineer on your team. Happy to keep this to a short call or async reply.

Best,
[Your Name]"""

    sections: list[tuple[str, str | None, bool]] = [
        ("EMAIL TEMPLATES — Cold Outreach", None, False),
        ("", None, False),
        ("VERSION A — Recruiters / Talent Acquisition", None, False),
        ("Short, direct ask. Use when contacting TA or recruiters.", None, False),
        (version_a, "email", True),
        ("", None, False),
        ("VERSION B — Engineering Managers (advice framing)", None, False),
        ("Ask for advice, not a job — often gets higher reply rates from EM/IC leads.", None, False),
        (version_b, "email", True),
        ("", None, False),
        ("WHY VERSION B WORKS", None, False),
        ("• Lower pressure: people help with advice more readily than 'give me a job'.", None, False),
        ("• Starts a conversation that can naturally lead to intros or open roles.", None, False),
        ("• Shows respect for their time and signals genuine interest in their team.", None, False),
        ("", None, False),
        ("EXAMPLE HOOKS (replace {{Hook}})", None, False),
        (
            "1. Razorpay: 'Razorpay's merchant onboarding flow is a benchmark for Indian fintech UX — that's the product bar I want to build toward.'",
            None,
            False,
        ),
        (
            "2. Yellow.ai: 'Yellow.ai's conversational AI stack sits at the intersection of product web and ML — exactly where I want to deepen my craft.'",
            None,
            False,
        ),
        (
            "3. Freshworks: 'Freshworks scaling SaaS from Chennai to global markets is the kind of product engineering story I want to be part of.'",
            None,
            False,
        ),
        (
            "4. Juspay: 'Juspay powering checkout for India's largest apps is the scale and fintech depth I'm targeting in my next role.'",
            None,
            False,
        ),
        (
            "5. GitLab: 'GitLab's all-remote engineering culture and DevOps platform align with how I already work — async, docs-first, web-heavy.'",
            None,
            False,
        ),
        ("", None, False),
        ("FOLLOW-UP (send 5 days after Version A or B if no reply)", None, False),
        (follow_up, "email", True),
    ]

    row = 1
    for text, kind, is_mono in sections:
        cell = ws.cell(row=row, column=1, value=text)
        if kind == "email":
            cell.font = MONO_FONT
            cell.fill = EMAIL_BLOCK_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            line_count = text.count("\n") + 1
            ws.row_dimensions[row].height = max(16 * line_count, 80)
        elif text.startswith("VERSION") or text.startswith("WHY") or text.startswith("EXAMPLE") or text.startswith("FOLLOW") or text.startswith("EMAIL"):
            cell.font = BOLD_FONT
        else:
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1


def main() -> None:
    with COMPANIES_FILE.open(encoding="utf-8") as f:
        companies = json.load(f)

    if len(companies) != 50:
        raise SystemExit(f"Expected 50 companies, got {len(companies)}")

    wb = Workbook()
    build_readme(wb)
    build_company_tracker(wb, companies)
    build_outreach_tracker(wb)
    build_email_template(wb)

    wb.save(OUTPUT)

    tier_counts: dict[int, int] = {}
    for co in companies:
        tier_counts[co["tier_key"]] = tier_counts.get(co["tier_key"], 0) + 1

    print(f"Saved {OUTPUT}")
    print(f"Companies: {len(companies)}")
    for tier in sorted(tier_counts):
        print(f"  Tier {tier}: {tier_counts[tier]}")
    print("Formula cells: Days Since Sent (101 rows), company/outreach stat counters")


if __name__ == "__main__":
    main()
