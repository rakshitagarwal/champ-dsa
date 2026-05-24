#!/usr/bin/env python3
"""Verify job_search_tracker.xlsx structure, formulas, and hyperlinks."""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

DIR = Path(__file__).resolve().parent
XLSX = DIR / "job_search_tracker.xlsx"

EXPECTED_SHEETS = ["README", "Company Tracker", "Outreach Tracker", "Email Template"]
FORMULA_ERRORS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!", "#N/A")
DAYS_SINCE_RE = re.compile(r'^=IF\(K\d+="","",TODAY\(\)-K\d+\)$')


def main() -> int:
    if not XLSX.exists():
        print(f"Missing file: {XLSX}")
        return 1

    wb = load_workbook(XLSX, data_only=False)
    errors: list[str] = []

    if wb.sheetnames != EXPECTED_SHEETS:
        errors.append(f"Sheet names: expected {EXPECTED_SHEETS}, got {wb.sheetnames}")

    # Company Tracker
    co = wb["Company Tracker"]
    if co["B1"].value != "=COUNTA(B5:B54)":
        errors.append(f"Company B1 formula: {co['B1'].value}")
    if co["B2"].value != "=SUM(J5:J54)":
        errors.append(f"Company B2 formula: {co['B2'].value}")
    if co["B3"].value != "=SUM(K5:K54)":
        errors.append(f"Company B3 formula: {co['B3'].value}")

    company_rows = [co.cell(row=r, column=2).value for r in range(5, 55)]
    if len([c for c in company_rows if c]) != 50:
        errors.append(f"Expected 50 companies, found {len([c for c in company_rows if c])}")

    for row in range(5, 55):
        li = co.cell(row=row, column=6)
        careers = co.cell(row=row, column=7)
        if not li.hyperlink or not li.hyperlink.target:
            errors.append(f"Row {row}: missing LinkedIn hyperlink")
        if not careers.hyperlink or not careers.hyperlink.target:
            errors.append(f"Row {row}: missing Careers hyperlink")
        if "linkedin.com/search/results/people" not in (li.hyperlink.target or ""):
            errors.append(f"Row {row}: LinkedIn URL malformed")

    if co.freeze_panes != "B5":
        errors.append(f"Company freeze_panes: {co.freeze_panes}")

    # Outreach Tracker
    out = wb["Outreach Tracker"]
    if out["B3"].value != "=COUNTA(K6:K106)":
        errors.append(f"Outreach B3 formula: {out['B3'].value}")
    if out["B4"].value != '=COUNTIF(N6:N106,"Yes")':
        errors.append(f"Outreach B4 formula: {out['B4'].value}")
    if out["E1"].value != '=IF(B3=0,"",B4/B3)':
        errors.append(f"Outreach reply rate formula: {out['E1'].value}")

    formula_rows = list(range(6, 107))
    bad_formulas = []
    for row in formula_rows:
        val = out.cell(row=row, column=12).value
        if not isinstance(val, str) or not DAYS_SINCE_RE.match(val):
            bad_formulas.append((row, val))
    if bad_formulas:
        errors.append(f"Bad Days Since Sent formulas: {bad_formulas[:5]}...")

    sample = out.cell(row=6, column=2).value
    if sample != "Priya":
        errors.append(f"Sample row expected at row 6, First Name={sample}")

    empty_count = sum(
        1 for r in range(7, 107) if not out.cell(row=r, column=2).value
    )
    if empty_count != 100:
        errors.append(f"Expected 100 empty outreach rows, found {empty_count}")

    if out.freeze_panes != "B6":
        errors.append(f"Outreach freeze_panes: {out.freeze_panes}")

    # Check for obvious formula error strings
    for ws_name in EXPECTED_SHEETS:
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in FORMULA_ERRORS:
                    errors.append(f"{ws_name}!{cell.coordinate}: {cell.value}")

    if errors:
        print("VERIFICATION FAILED:")
        for e in errors:
            print(f"  - {e}")
        return 1

    print(f"OK: {XLSX}")
    print("  4 sheets, 50 companies, 101 outreach rows (1 sample + 100 empty)")
    print("  Hyperlinks, formulas, and freeze panes verified")
    return 0


if __name__ == "__main__":
    sys.exit(main())
